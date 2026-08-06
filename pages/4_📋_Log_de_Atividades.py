"""
Log de Atividades — o que a equipe fez no painel.

Analista vê o próprio log; coordenador e gerente veem a equipe abaixo deles;
diretor vê tudo. Esse filtro é aplicado pelo banco (RLS), não aqui — então não
tem como alguém ver mais do que deveria mexendo na tela.
"""
import io
from datetime import date, timedelta

import pandas as pd
import streamlit as st

from core.auth import exigir_login, cargo_do_usuario
from core.auditoria import buscar, ACOES

st.set_page_config(page_title="Log de Atividades — Painel DP", page_icon="📋", layout="wide")
exigir_login()

st.title("📋 Log de Atividades")

cargo = cargo_do_usuario()
if cargo == "diretor":
    st.caption("Você é diretora: está vendo o que **toda a equipe** fez no painel.")
elif cargo in ("gerente", "coordenador"):
    st.caption("Você está vendo o que **você e a sua equipe** fizeram no painel.")
else:
    st.caption("Você está vendo o seu próprio histórico de uso do painel.")

try:
    registros = buscar()
except Exception as e:
    st.error(
        "Não consegui ler o log. Se a mensagem falar em 'log_atividades', é sinal de que "
        f"o script db/fix_e_log.sql ainda não foi rodado no Supabase.\n\nDetalhe: {e}"
    )
    st.stop()

if not registros:
    st.info(
        "Nenhuma atividade registrada ainda. O log começa a se preencher sozinho conforme "
        "a equipe usa o painel (login, conferências, cadastros, alterações de parâmetros)."
    )
    st.stop()

df = pd.DataFrame(registros)
df["criado_em"] = pd.to_datetime(df["criado_em"], errors="coerce", utc=True)
try:
    df["criado_em"] = df["criado_em"].dt.tz_convert("America/Sao_Paulo")
except Exception:
    pass

# ------------------------------------------------------------------ filtros
c1, c2, c3 = st.columns([2, 2, 3])

pessoas = sorted({(n or "(sem nome)") for n in df["usuario_nome"].fillna("")})
pessoa_sel = c1.multiselect("Pessoa", options=pessoas, default=[])

acoes_presentes = [a for a in ACOES if a in set(df["acao"])]
acoes_extra = sorted(set(df["acao"]) - set(ACOES))
acao_sel = c2.multiselect("Ação", options=acoes_presentes + acoes_extra, default=[])

hoje = date.today()
periodo = c3.date_input(
    "Período",
    value=(hoje - timedelta(days=30), hoje),
    format="DD/MM/YYYY",
)

filtrado = df.copy()
if pessoa_sel:
    filtrado = filtrado[filtrado["usuario_nome"].fillna("(sem nome)").isin(pessoa_sel)]
if acao_sel:
    filtrado = filtrado[filtrado["acao"].isin(acao_sel)]
if isinstance(periodo, (tuple, list)) and len(periodo) == 2:
    ini, fim = periodo
    somente_data = filtrado["criado_em"].dt.date
    filtrado = filtrado[(somente_data >= ini) & (somente_data <= fim)]

# ------------------------------------------------------------------ resumo
m1, m2, m3, m4 = st.columns(4)
m1.metric("Atividades no período", len(filtrado))
m2.metric("Pessoas ativas", filtrado["usuario_nome"].nunique())
m3.metric("Conferências de férias", int((filtrado["acao"] == "Conferência de férias").sum()))
m4.metric("Clientes envolvidos", filtrado["empresa_nome"].replace("", pd.NA).nunique())

st.divider()

tabela = pd.DataFrame({
    "Data/hora": filtrado["criado_em"].dt.strftime("%d/%m/%Y %H:%M"),
    "Pessoa": filtrado["usuario_nome"].replace("", "(sem nome)"),
    "Cargo": filtrado["cargo"].fillna("").str.capitalize(),
    "Ação": filtrado["acao"],
    "Cliente": filtrado["empresa_nome"].fillna(""),
    "Detalhe": filtrado["detalhe"].fillna(""),
    "E-mail": filtrado["usuario_email"].fillna(""),
})

st.dataframe(tabela, use_container_width=True, height=520, hide_index=True)

# ------------------------------------------------------------------ export
buf = io.BytesIO()
with pd.ExcelWriter(buf, engine="openpyxl") as writer:
    tabela.to_excel(writer, index=False, sheet_name="Log")
    ws = writer.sheets["Log"]
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    for c in range(1, len(tabela.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions[get_column_letter(c)].width = 22

st.download_button(
    "⬇️ Baixar log em Excel",
    data=buf.getvalue(),
    file_name="Log_Atividades_Painel_DP.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.caption(
    "O log é imutável: ninguém — nem a diretoria — consegue editar ou apagar registros, "
    "porque o banco só permite inserção. Cada pessoa só grava em seu próprio nome."
)
