"""
Motor de cálculo — Conferência de Férias

Lê um workbook mensal de "Programação de Férias" (exportado do sistema de
folha, abas 'Todas', '644', 'FICHA' e opcionalmente 'FÉRIAS CADASTRADAS
DEPOIS') e recalcula, funcionário a funcionário, o líquido de férias de
forma independente do sistema — mesma lógica usada na aba manual "CONFERE"
que a Sinergy já usava, só que parametrizada (tabelas de INSS/IRRF editáveis)
e reutilizável para qualquer cliente que exporte no mesmo formato.

Este módulo não depende de Streamlit nem de Supabase — só de openpyxl e
pandas — para poder ser testado isoladamente.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
import pandas as pd

REQUIRED_SHEETS = ["Todas", "644", "FICHA"]
DEPOIS_SHEET = "FÉRIAS CADASTRADAS DEPOIS"


class FormatoInvalido(Exception):
    """Levantado quando o arquivo enviado não tem as abas esperadas."""


@dataclass
class Parametros:
    """Tabelas de INSS/IRRF e deduções. Cada faixa é (valor_ate, aliquota, parcela_deduzir).
    A última faixa de cada tabela é aberta (sem limite superior / valor_ate=None
    para IRRF; para INSS o valor_ate da última faixa funciona como teto)."""

    inss: list = field(default_factory=lambda: [
        (1621.01, 0.075, 0.0),
        (2901.84, 0.09, 24.32),
        (4354.27, 0.12, 111.40),
        (8475.55, 0.14, 198.49),  # última = teto de contribuição
    ])
    irrf: list = field(default_factory=lambda: [
        (2428.80, 0.0, 0.0),
        (2826.65, 0.075, 182.16),
        (3751.05, 0.15, 394.16),
        (4664.68, 0.225, 675.49),
        (None, 0.275, 908.73),  # faixa aberta (sem limite superior)
    ])
    dep_deducao: float = 189.59
    ded_simplificada: float = 607.20
    redutor_limite: float = 7350.00
    redutor_a: float = 978.62
    redutor_b: float = 0.133145
    tolerancia: float = 0.05

    def calcular_inss(self, base: float) -> float:
        for ate, aliq, ded in self.inss[:-1]:
            if base < ate:
                return base * aliq - ded
        ate, aliq, ded = self.inss[-1]
        if base < ate:
            return base * aliq - ded
        return ate * aliq - ded  # aplica o teto

    def calcular_irrf(self, base: float) -> float:
        for ate, aliq, ded in self.irrf[:-1]:
            if ate is not None and base < ate:
                return base * aliq - ded
        aliq, ded = self.irrf[-1][1], self.irrf[-1][2]
        return base * aliq - ded

    def calcular_redutor(self, base_bruta_ir: float) -> float:
        if base_bruta_ir < self.redutor_limite + 0.01:
            return self.redutor_a - (base_bruta_ir * self.redutor_b)
        return 0.0


def _find_header_row(ws, header_text: str, key_col: int = 1) -> int:
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=key_col).value
        if isinstance(val, str) and val.strip().upper() == header_text.upper():
            return r
    raise FormatoInvalido(f"Não encontrei o cabeçalho '{header_text}' na aba '{ws.title}'")


def _read_employee_rows(ws, header_text: str = "MAT", key_col: int = 1):
    """Lista (linha, matrícula) para linhas de funcionário de verdade, pulando
    linhas de subcabeçalho de filial/centro de custo (texto na coluna MAT)."""
    header_row = _find_header_row(ws, header_text, key_col=key_col)
    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        mat = ws.cell(row=r, column=key_col).value
        nome = ws.cell(row=r, column=key_col + 1).value
        if mat is None and (nome is None or not str(nome).strip()):
            break
        if isinstance(mat, str) and mat.strip().lower().startswith("total"):
            break
        if isinstance(mat, (int, float)):
            rows.append((r, int(mat)))
        r += 1
    return rows


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _build_employee_plan(wb):
    ws_todas = wb["Todas"]
    todas_rows = _read_employee_rows(ws_todas)
    plan = {mat: {"mat": mat, "sheet": "Todas", "row": r, "origem": "Programação"} for r, mat in todas_rows}
    order = [mat for _, mat in todas_rows]

    if DEPOIS_SHEET in wb.sheetnames:
        ws_depois = wb[DEPOIS_SHEET]
        for r, mat in _read_employee_rows(ws_depois):
            if mat in plan:
                plan[mat] = {"mat": mat, "sheet": DEPOIS_SHEET, "row": r, "origem": "Cadastrada depois (atualizada)"}
            else:
                plan[mat] = {"mat": mat, "sheet": DEPOIS_SHEET, "row": r, "origem": "Cadastrada depois (nova)"}
                order.append(mat)

    return [plan[mat] for mat in order]


def _build_ficha_index(wb):
    """FICHA é uma tabela dinâmica (valores prontos, sem fórmula). Colunas fixas
    conforme o layout padrão do sistema: A=MAT, Q=dependentes, T=eConsignado,
    Y=líquido, AA=média de férias (1-indexado: Q=17, T=20, Y=25, AA=27)."""
    ws = wb["FICHA"]
    idx = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        mat = row[0]
        if not isinstance(mat, (int, float)):
            continue
        mat = int(mat)
        idx[mat] = {
            "dependentes": _num(row[16]) if len(row) > 16 else 0.0,   # col Q (17ª)
            "econsignado": _num(row[19]) if len(row) > 19 else 0.0,   # col T (20ª)
            "liquido_sistema": _num(row[24]) if len(row) > 24 else 0.0,  # col Y (25ª)
            "media_ferias": _num(row[26]) if len(row) > 26 else 0.0,  # col AA (27ª)
        }
    return idx


def _build_644_index(wb):
    ws = wb["644"]
    idx = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat = row[0]
        if not isinstance(mat, (int, float)):
            continue
        idx[int(mat)] = _num(row[7]) if len(row) > 7 else 0.0  # coluna H = valor
    return idx


def carregar_workbook(conteudo: bytes):
    """Abre o workbook em memória e valida se tem as abas obrigatórias."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    faltando = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if faltando:
        raise FormatoInvalido(
            "Aba(s) obrigatória(s) não encontrada(s) no arquivo: " + ", ".join(faltando)
        )
    return wb


def calcular_conferencia(wb, parametros: Optional[Parametros] = None) -> pd.DataFrame:
    """Recalcula o líquido de férias de cada funcionário e retorna um DataFrame
    já pronto para exibir/exportar, com STATUS = OK / VERIFICAR / SEM DADOS NO SISTEMA."""
    p = parametros or Parametros()
    ws_todas = wb["Todas"]

    plano = _build_employee_plan(wb)
    ficha_idx = _build_ficha_index(wb)
    idx_644 = _build_644_index(wb)

    linhas = []
    for emp in plano:
        mat = emp["mat"]
        ws_src = wb[emp["sheet"]]
        t = emp["row"]

        nome = ws_src.cell(row=t, column=2).value
        data_inicio = ws_src.cell(row=t, column=11).value   # K
        data_fim = ws_src.cell(row=t, column=12).value      # L
        data_credito = ws_src.cell(row=t, column=13).value  # M
        dias = _num(ws_src.cell(row=t, column=14).value)    # N
        dias_abono = _num(ws_src.cell(row=t, column=15).value)  # O
        sal13 = ws_src.cell(row=t, column=16).value          # P

        tem_dados = mat in ficha_idx
        ficha = ficha_idx.get(mat, {})
        dep = ficha.get("dependentes", 0.0)
        liquido_sistema = ficha.get("liquido_sistema", 0.0)
        media_ferias = ficha.get("media_ferias", 0.0)
        econsignado = ficha.get("econsignado", 0.0)
        salario_base = idx_644.get(mat, 0.0)

        ferias = salario_base / 30 * dias
        terco_media_ferias = media_ferias / 3
        terco_ferias = ferias / 3
        abono = salario_base / 30 * dias_abono
        terco_abono = abono / 3
        media_abono = (media_ferias / dias * dias_abono) if dias else 0.0
        terco_media_abono = media_abono / 3
        decimo_terceiro = salario_base / 2 if str(sal13).strip().upper() == "SIM" else 0.0

        base_inss = ferias + media_ferias + terco_media_ferias + terco_ferias
        inss = p.calcular_inss(base_inss)
        base_bruta_ir = base_inss + terco_abono + terco_media_abono
        deducao_dependente = dep * p.dep_deducao if dep > 0 else 0.0
        if (inss + deducao_dependente) < p.ded_simplificada:
            base_ir_com_deducao = base_bruta_ir - p.ded_simplificada
        else:
            base_ir_com_deducao = base_bruta_ir - (inss + deducao_dependente)
        irrf_calc01 = p.calcular_irrf(base_ir_com_deducao)
        irrf_calc02 = p.calcular_redutor(base_bruta_ir)
        irrf_desconto = max(irrf_calc01 - irrf_calc02, 0.0)
        pensao_manual = 0.0

        liquido_recalculado = (
            ferias + media_ferias + terco_media_ferias + terco_ferias
            + abono + terco_abono + media_abono + terco_media_abono + decimo_terceiro
            - inss - irrf_desconto - pensao_manual - econsignado
        )
        diferenca = liquido_sistema - liquido_recalculado

        if not tem_dados:
            status = "SEM DADOS NO SISTEMA"
        elif abs(diferenca) <= p.tolerancia:
            status = "OK"
        else:
            status = "VERIFICAR"

        linhas.append({
            "MAT": mat,
            "NOME": nome,
            "DATA INÍCIO": data_inicio,
            "DATA FIM": data_fim,
            "DATA CRÉDITO": data_credito,
            "DIAS": dias,
            "DIAS ABONO": dias_abono,
            "13º SAL.": sal13,
            "ORIGEM": emp["origem"],
            "DADOS NO SISTEMA?": "SIM" if tem_dados else "NÃO",
            "DEP. IR": dep,
            "LÍQUIDO (SISTEMA)": round(liquido_sistema, 2),
            "SALÁRIO BASE": round(salario_base, 2),
            "INSS": round(inss, 2),
            "IRRF": round(irrf_desconto, 2),
            "PENSÃO (MANUAL)": pensao_manual,
            "eCONSIGNADO": round(econsignado, 2),
            "LÍQUIDO (RECALCULADO)": round(liquido_recalculado, 2),
            "DIFERENÇA": round(diferenca, 2),
            "STATUS": status,
        })

    return pd.DataFrame(linhas)


def resumo(df: pd.DataFrame) -> dict:
    total = len(df)
    ok = int((df["STATUS"] == "OK").sum())
    verificar = int((df["STATUS"] == "VERIFICAR").sum())
    sem_dados = int((df["STATUS"] == "SEM DADOS NO SISTEMA").sum())
    return {
        "total_funcionarios": total,
        "ok": ok,
        "verificar": verificar,
        "sem_dados": sem_dados,
        "soma_liquido_sistema": round(float(df["LÍQUIDO (SISTEMA)"].sum()), 2),
        "soma_liquido_recalculado": round(float(df["LÍQUIDO (RECALCULADO)"].sum()), 2),
    }


def gerar_excel_bytes(df: pd.DataFrame) -> bytes:
    """Gera um .xlsx (valores estáticos, com formatação condicional simples) para download."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Conferência")
        wsx = writer.sheets["Conferência"]
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ok_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        alert_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        for c in range(1, len(df.columns) + 1):
            cell = wsx.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            wsx.column_dimensions[get_column_letter(c)].width = 16

        status_col = df.columns.get_loc("STATUS") + 1
        for r in range(2, len(df) + 2):
            status = wsx.cell(row=r, column=status_col).value
            fill = {"OK": ok_fill, "VERIFICAR": alert_fill, "SEM DADOS NO SISTEMA": warn_fill}.get(status)
            if fill:
                for c in range(1, len(df.columns) + 1):
                    wsx.cell(row=r, column=c).fill = fill

    return buf.getvalue()
