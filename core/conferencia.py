"""
Conferência de Férias a partir dos arquivos crus — monta o "de-para".

Junta as três fontes já tratadas em core/preparacao.py e produz uma linha por
funcionário com, lado a lado:
  * o que o SISTEMA calculou (vindo da Consulta Ficha do mês das férias);
  * o que o painel RECALCULOU de forma independente, a partir do salário da
    verba 644 do mês anterior e das tabelas de INSS/IRRF configuradas;
  * a diferença e um STATUS.

O layout das colunas segue a planilha CONFERE que a Sinergy já usava, para a
leitura ficar familiar.

Regra de negócio que vale registrar: as férias de um mês são calculadas sobre o
salário do MÊS ANTERIOR (férias de setembro usam o salário de agosto). Por isso
a ficha de salário pedida é sempre a da competência anterior.
"""
from __future__ import annotations

import io

import pandas as pd

from core.calculo import Parametros

COLUNAS = [
    "MAT", "NOME", "DATA INÍCIO", "DATA FIM", "DATA CRÉDITO", "DIAS", "DIAS ABONO",
    "13º SAL.", "COMP. CRÉDITO", "DEP. IR", "644 (SALÁRIO)",
    "FÉRIAS", "MÉDIA FÉRIAS", "1/3 MÉDIA FÉRIAS", "1/3 FÉRIAS",
    "ABONO", "1/3 ABONO", "MÉDIA ABONO", "1/3 MÉDIA ABONO", "13º",
    "BASE INSS", "INSS", "BASE BRUTA IR", "DEDUÇÃO DEPENDENTE", "BASE IR C/ DEDUÇÃO",
    "IRRF CALC 01", "IRRF CALC 02", "IRRF DESCONTO", "eCONSIGNADO", "PENSÃO",
    "LÍQUIDO (SISTEMA)", "LÍQUIDO (RECALCULADO)", "DIFERENÇA", "STATUS", "OBSERVAÇÃO",
]


def _competencia_da_data(d) -> str:
    return f"{d.month:02d}/{d.year}" if d else ""


def montar_conferencia(programacao: pd.DataFrame, salarios: dict, ficha: dict,
                       parametros: Parametros | None = None,
                       competencia_ferias: str = "") -> pd.DataFrame:
    """programacao: saída de preparacao.ler_programacao
    salarios: matrícula -> salário 644 do mês anterior
    ficha: matrícula -> verbas do mês das férias (o cálculo do sistema)
    competencia_ferias: '09/2026' — usada para saber de quem é o recibo deste mês

    Regra importante: o recibo de férias pertence à competência da DATA CRÉDITO
    (a data em que o valor é pago), não à data de início das férias. Quem sai de
    férias em setembro mas recebeu em agosto tem o recibo na ficha de agosto —
    por isso essas pessoas não são divergência, são conferência de outro mês."""
    p = parametros or Parametros()
    competencia_salario_txt = ""
    linhas = []

    for _, f in programacao.iterrows():
        mat = int(f["MAT"])
        sistema = ficha.get(mat, {})
        salario = float(salarios.get(mat, 0.0))
        tem_sistema = mat in ficha
        tem_salario = mat in salarios

        dias = float(f["DIAS"] or 0)
        dias_abono = float(f["DIAS ABONO"] or 0)

        # valores que o painel não tem como deduzir sozinho e por isso vêm do
        # sistema: médias (dependem do histórico variável), dependentes de IR,
        # eConsignado e pensão alimentícia (essa é sempre manual)
        media_ferias = float(sistema.get("media_ferias", 0.0))
        media_abono = float(sistema.get("media_abono", 0.0))
        dependentes = float(sistema.get("dependentes", 0.0))
        econsignado = float(sistema.get("econsignado", 0.0))
        pensao = 0.0

        # o recibo é do mês em que o valor foi creditado
        comp_credito = _competencia_da_data(f["DATA CRÉDITO"])
        recibo_neste_mes = (not competencia_ferias) or (comp_credito == competencia_ferias)

        # salário que o sistema realmente usou no cálculo das férias
        salario_usado_sistema = float(sistema.get("salario_644", 0.0))

        # ------------------------------------------------ recálculo independente
        ferias = salario / 30 * dias
        terco_ferias = ferias / 3
        terco_media_ferias = media_ferias / 3
        abono = salario / 30 * dias_abono
        terco_abono = abono / 3
        terco_media_abono = media_abono / 3
        # O adiantamento de 13º não é recalculado: depende de avos já pagos no ano,
        # que não vêm em nenhum destes arquivos. Usamos o valor do próprio sistema
        # (verbas 80 / 494) e sinalizamos quando a programação pede e ele não veio.
        pediu_13 = str(f["13º SAL."]).strip().upper() == "SIM"
        decimo = float(sistema.get("decimo_terceiro", 0.0))

        base_inss = ferias + media_ferias + terco_ferias + terco_media_ferias
        inss = p.calcular_inss(base_inss) if base_inss > 0 else 0.0

        base_bruta_ir = base_inss + terco_abono + terco_media_abono
        deducao_dependente = dependentes * p.dep_deducao if dependentes > 0 else 0.0
        if (inss + deducao_dependente) < p.ded_simplificada:
            base_ir = base_bruta_ir - p.ded_simplificada
        else:
            base_ir = base_bruta_ir - (inss + deducao_dependente)
        base_ir = max(base_ir, 0.0)

        irrf_01 = p.calcular_irrf(base_ir) if base_ir > 0 else 0.0
        irrf_02 = p.calcular_redutor(base_bruta_ir)
        irrf = max(irrf_01 - irrf_02, 0.0)

        liquido_calc = (
            ferias + media_ferias + terco_ferias + terco_media_ferias
            + abono + terco_abono + media_abono + terco_media_abono + decimo
            - inss - irrf - pensao - econsignado
        )
        liquido_sistema = float(sistema.get("liquido", 0.0))
        diferenca = liquido_sistema - liquido_calc

        # ------------------------------------------------------------- status
        obs = []
        inss_sistema = float(sistema.get("inss", 0.0))
        irrf_sistema = float(sistema.get("irrf", 0.0))

        # o sistema calculou as férias sobre um salário diferente do da folha?
        if tem_salario and salario_usado_sistema and abs(salario_usado_sistema - salario) > 0.01:
            obs.append(
                f"ATENÇÃO: a folha do mês anterior traz salário R$ {salario:,.2f}, mas as férias "
                f"foram calculadas sobre R$ {salario_usado_sistema:,.2f}."
            )

        if not tem_salario:
            status = "SEM SALÁRIO"
            obs.append("Matrícula sem verba 644 na ficha do mês anterior — confira se foi "
                       "admitida depois do fechamento.")
        elif not recibo_neste_mes:
            status = "RECIBO EM OUTRA COMPETÊNCIA"
            data_cred = f["DATA CRÉDITO"].strftime("%d/%m/%Y") if f["DATA CRÉDITO"] else "—"
            obs.append(
                f"Crédito em {data_cred}: o recibo foi pago na competência {comp_credito}, "
                f"não em {competencia_ferias}. Confira esta pessoa junto com a ficha daquele mês."
            )
            liquido_sistema = 0.0
            diferenca = 0.0
        elif not tem_sistema:
            status = "SEM CÁLCULO NO SISTEMA"
            obs.append("Está na programação de férias mas o sistema ainda não calculou.")
        elif abs(diferenca) <= p.tolerancia:
            status = "OK"
        else:
            status = "VERIFICAR"
            if econsignado and abs(abs(diferenca) - econsignado) <= p.tolerancia:
                obs.append("A diferença é exatamente o eConsignado.")
            elif inss_sistema and abs(inss_sistema - inss) > p.tolerancia:
                obs.append(f"INSS: sistema R$ {inss_sistema:,.2f} x recalculado R$ {inss:,.2f}.")
            elif irrf_sistema and abs(irrf_sistema - irrf) > p.tolerancia:
                ded_legal = float(sistema.get("deducao_legal", 0.0))
                ded_simpl = float(sistema.get("deducao_simplificada", 0.0))
                usada = ded_legal or ded_simpl
                cabivel = max(inss + deducao_dependente, p.ded_simplificada)
                obs.append(
                    f"IRRF: sistema R$ {irrf_sistema:,.2f} x recalculado R$ {irrf:,.2f}. O sistema "
                    f"deduziu R$ {usada:,.2f} ({'legal' if ded_legal else 'simplificada'}); pela regra "
                    f"da dedução mais vantajosa caberia R$ {cabivel:,.2f}."
                )
            else:
                obs.append("Verifique pensão alimentícia (sempre manual) ou as médias.")

        if pediu_13 and decimo == 0 and recibo_neste_mes and tem_sistema:
            obs.append("A programação pede adiantamento de 13º, mas o sistema não lançou a verba.")

        linhas.append({
            "MAT": mat,
            "NOME": f["NOME"],
            "DATA INÍCIO": f["DATA INÍCIO"],
            "DATA FIM": f["DATA FIM"],
            "DATA CRÉDITO": f["DATA CRÉDITO"],
            "DIAS": dias,
            "DIAS ABONO": dias_abono,
            "13º SAL.": f["13º SAL."],
            "COMP. CRÉDITO": comp_credito,
            "DEP. IR": dependentes,
            "644 (SALÁRIO)": round(salario, 2),
            "FÉRIAS": round(ferias, 2),
            "MÉDIA FÉRIAS": round(media_ferias, 2),
            "1/3 MÉDIA FÉRIAS": round(terco_media_ferias, 2),
            "1/3 FÉRIAS": round(terco_ferias, 2),
            "ABONO": round(abono, 2),
            "1/3 ABONO": round(terco_abono, 2),
            "MÉDIA ABONO": round(media_abono, 2),
            "1/3 MÉDIA ABONO": round(terco_media_abono, 2),
            "13º": round(decimo, 2),
            "BASE INSS": round(base_inss, 2),
            "INSS": round(inss, 2),
            "BASE BRUTA IR": round(base_bruta_ir, 2),
            "DEDUÇÃO DEPENDENTE": round(deducao_dependente, 2),
            "BASE IR C/ DEDUÇÃO": round(base_ir, 2),
            "IRRF CALC 01": round(irrf_01, 2),
            "IRRF CALC 02": round(irrf_02, 2),
            "IRRF DESCONTO": round(irrf, 2),
            "eCONSIGNADO": round(econsignado, 2),
            "PENSÃO": round(pensao, 2),
            "LÍQUIDO (SISTEMA)": round(liquido_sistema, 2),
            "LÍQUIDO (RECALCULADO)": round(liquido_calc, 2),
            "DIFERENÇA": round(diferenca, 2),
            "STATUS": status,
            "OBSERVAÇÃO": " ".join(obs),
        })

    return pd.DataFrame(linhas, columns=COLUNAS)


def resumo(df: pd.DataFrame) -> dict:
    return {
        "total_funcionarios": len(df),
        "ok": int((df["STATUS"] == "OK").sum()),
        "verificar": int((df["STATUS"] == "VERIFICAR").sum()),
        "sem_dados": int(df["STATUS"].isin(["SEM SALÁRIO", "SEM CÁLCULO NO SISTEMA"]).sum()),
        "soma_liquido_sistema": round(float(df["LÍQUIDO (SISTEMA)"].sum()), 2),
        "soma_liquido_recalculado": round(float(df["LÍQUIDO (RECALCULADO)"].sum()), 2),
        "diferenca_total": round(float(df["DIFERENÇA"].sum()), 2),
    }


CORES = {
    "OK": "DCFCE7",
    "VERIFICAR": "FEE2E2",
    "SEM SALÁRIO": "FEF3C7",
    "SEM CÁLCULO NO SISTEMA": "FEF3C7",
}


def gerar_excel(df: pd.DataFrame, competencia_ferias: str = "",
                competencia_salario: str = "") -> bytes:
    """Gera o Excel do de-para, com uma aba de conferência e uma de resumo."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Conferência", startrow=2)
        ws = writer.sheets["Conferência"]

        titulo = "Conferência de Férias"
        if competencia_ferias:
            titulo += f" — competência {competencia_ferias}"
        if competencia_salario:
            titulo += f" (salário base: {competencia_salario})"
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13)

        cabecalho = PatternFill("solid", start_color="1F2937", end_color="1F2937")
        for c in range(1, len(df.columns) + 1):
            cel = ws.cell(row=3, column=c)
            cel.fill = cabecalho
            cel.font = Font(color="FFFFFF", bold=True)
            cel.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            largura = 30 if df.columns[c - 1] in ("NOME", "OBSERVAÇÃO") else 15
            ws.column_dimensions[get_column_letter(c)].width = largura
        ws.row_dimensions[3].height = 32

        col_status = list(df.columns).index("STATUS") + 1
        for r in range(4, len(df) + 4):
            cor = CORES.get(ws.cell(row=r, column=col_status).value)
            if cor:
                fill = PatternFill("solid", start_color=cor, end_color=cor)
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=r, column=c).fill = fill
            for c in range(1, len(df.columns) + 1):
                valor = ws.cell(row=r, column=c).value
                if isinstance(valor, float):
                    ws.cell(row=r, column=c).number_format = "#,##0.00"
        ws.freeze_panes = "C4"

        r = resumo(df)
        pd.DataFrame([
            {"Indicador": "Funcionários conferidos", "Valor": r["total_funcionarios"]},
            {"Indicador": "OK", "Valor": r["ok"]},
            {"Indicador": "A verificar", "Valor": r["verificar"]},
            {"Indicador": "Sem dados suficientes", "Valor": r["sem_dados"]},
            {"Indicador": "Líquido somado (sistema)", "Valor": r["soma_liquido_sistema"]},
            {"Indicador": "Líquido somado (recalculado)", "Valor": r["soma_liquido_recalculado"]},
            {"Indicador": "Diferença total", "Valor": r["diferenca_total"]},
            {"Indicador": "Competência das férias", "Valor": competencia_ferias},
            {"Indicador": "Competência do salário (644)", "Valor": competencia_salario},
        ]).to_excel(writer, index=False, sheet_name="Resumo")
        ws2 = writer.sheets["Resumo"]
        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 22
        for c in (1, 2):
            cel = ws2.cell(row=1, column=c)
            cel.fill = cabecalho
            cel.font = Font(color="FFFFFF", bold=True)

    return buf.getvalue()
