"""
Tratamento dos arquivos CRUS exportados do sistema Sinergy.

Antes o painel só aceitava uma planilha já montada à mão (com as abas 'Todas',
'644' e 'FICHA' prontas). Agora ele recebe os arquivos exatamente como saem do
sistema e monta tudo sozinho:

  1) Relatório de Programação de Férias  -> quem entra de férias, datas, dias,
     dias de abono e se adianta o 13º. É a BASE: quem não está aqui não entra
     na conferência.
  2) Consulta Ficha do MÊS ANTERIOR      -> verba 644 (Salário Mês c/ Adic.),
     que é o salário usado para calcular as férias. Férias de setembro usam o
     salário de agosto.
  3) Consulta Ficha do MÊS DAS FÉRIAS    -> o cálculo que o sistema já fez
     (uma linha por verba). É contra isso que a conferência compara.

Os arquivos 2 e 3 vêm no formato "longo": uma linha por funcionário por verba.
Aqui eles viram uma linha por funcionário, com cada verba na sua coluna.

Não depende de Streamlit nem de Supabase — dá para testar sozinho.
"""
from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime

import openpyxl
import pandas as pd


class FormatoInvalido(Exception):
    """Arquivo enviado não tem a cara esperada."""


# --------------------------------------------------------------------------
# Verbas do sistema. O sufixo (REC) é o recibo de férias fechado — é o que a
# conferência usa. As versões "no Mês" / "Mês Seguinte" são o rateio contábil
# entre as competências e servem só de referência.
# --------------------------------------------------------------------------
VERBA_SALARIO = 644          # Salário Mês c/ Adic.

VERBAS = {
    "ferias":            (147, 148, 386),
    "media_ferias":      (144, 488, 387),
    "terco_ferias":      (139, 140, 388),
    "terco_media_ferias": (471, 482, 473),
    "abono":             (142, 474),
    "terco_abono":       (150, 475),
    "media_abono":       (143, 476),
    "terco_media_abono": (470, 477),
    "decimo_terceiro":   (80, 494),
    "base_inss":         (278,),
    "inss":              (199, 484, 389),
    "base_irrf":         (280,),
    "base_ir_com_deducao": (1953,),
    "irrf_provisorio":   (1954,),
    "redutor_irrf":      (1955,),
    "irrf":              (257, 408, 390),
    "liquido":           (439,),
    "dependentes":       (527,),
    "deducao_dependente": (14318, 14319, 14320),
    "deducao_simplificada": (871,),
    "deducao_legal":     (779,),
    "econsignado":       (1414, 1415, 1416, 1434),
    "desconto_liquido":  (261, 395),
    "recesso_estagiario": (1584, 1585),
}

# INSS por faixa progressiva — o sistema lança uma verba por faixa
VERBAS_INSS_FAIXAS = (1684, 1685, 1686, 1687)


def _texto(v) -> str:
    return "" if v is None else str(v).strip()


def _sem_acento(t: str) -> str:
    """Tira acento e normaliza os ordinais º/ª. Sem isso o cabeçalho '13º Sal.'
    não casava com nada e o adiantamento de 13º saía sempre como NÃO."""
    t = t.replace("º", "o").replace("ª", "a").replace("°", "o")
    return "".join(c for c in unicodedata.normalize("NFD", t)
                   if unicodedata.category(c) != "Mn").lower()


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in t:                       # 1.234,56 -> 1234.56
        t = t.replace(".", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def _data(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    t = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(t, fmt).date()
        except ValueError:
            continue
    return None


def _matricula(v):
    """Matrícula pode vir como número, texto ou com zeros à esquerda."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    t = re.sub(r"\D", "", str(v))
    return int(t) if t else None


# ==========================================================================
# 1) Relatório de Programação de Férias
# ==========================================================================
def ler_programacao(conteudo: bytes) -> pd.DataFrame:
    """Lê o relatório cru de Programação de Férias.

    O arquivo tem uma aba por empresa/filial e normalmente uma aba 'Todas' com
    o consolidado. Usamos 'Todas' quando existe; se não existir, juntamos todas
    as abas e removemos matrículas repetidas.

    Dentro da aba há linhas de subtítulo de centro de custo (ex.: 'BR49811000 -
    CODIR') sem matrícula — essas são ignoradas."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)

    abas = ["Todas"] if "Todas" in wb.sheetnames else list(wb.sheetnames)
    linhas, vistas = [], set()

    for nome in abas:
        ws = wb[nome]
        cab = _localizar_cabecalho(ws)
        if cab is None:
            continue
        linha_cab, col = cab
        for r in range(linha_cab + 1, ws.max_row + 1):
            mat = _matricula(ws.cell(row=r, column=col["MAT"]).value)
            if mat is None or mat in vistas:
                continue
            nome_func = _texto(ws.cell(row=r, column=col["NOME"]).value)
            if not nome_func:
                continue
            vistas.add(mat)
            linhas.append({
                "MAT": mat,
                "NOME": nome_func,
                "ADMISSÃO": _data(ws.cell(row=r, column=col.get("ADMISSAO", 3)).value),
                "DATA INÍCIO": _data(ws.cell(row=r, column=col["INICIO_FERIAS"]).value),
                "DATA FIM": _data(ws.cell(row=r, column=col["FIM_FERIAS"]).value),
                "DATA CRÉDITO": _data(ws.cell(row=r, column=col["CREDITO"]).value),
                "DIAS": _num(ws.cell(row=r, column=col["DIAS"]).value),
                "DIAS ABONO": _num(ws.cell(row=r, column=col["ABONO"]).value) if col.get("ABONO") else 0.0,
                "13º SAL.": _texto(ws.cell(row=r, column=col["DEC13"]).value).upper() if col.get("DEC13") else "NAO",
                "SITUAÇÃO": _texto(ws.cell(row=r, column=col["SITUACAO"]).value) if col.get("SITUACAO") else "",
                "ABA": nome,
            })

    if not linhas:
        raise FormatoInvalido(
            "Não encontrei nenhum funcionário no relatório de Programação de Férias. "
            "Confira se o arquivo é o export correto do sistema (ele precisa ter a "
            "coluna 'MAT' e 'NOME DO FUNCIONÁRIO')."
        )
    return pd.DataFrame(linhas)


def _localizar_cabecalho(ws):
    """Acha a linha do cabeçalho e mapeia as colunas pelo nome, em vez de assumir
    posições fixas — assim o painel não quebra se o sistema mudar a ordem."""
    for r in range(1, min(ws.max_row, 30) + 1):
        valores = {}
        for c in range(1, ws.max_column + 1):
            t = _sem_acento(_texto(ws.cell(row=r, column=c).value))
            if t:
                valores.setdefault(t, []).append(c)
        if "mat" not in valores:
            continue

        col = {"MAT": valores["mat"][0]}
        for chave, alvo in [("NOME", "nome do funcionario"), ("ADMISSAO", "admissao")]:
            if alvo in valores:
                col[chave] = valores[alvo][0]
        col.setdefault("NOME", col["MAT"] + 1)

        # 'DATA INÍCIO' e 'DATA FIM' aparecem duas vezes: período aquisitivo e
        # férias. A segunda ocorrência é a das férias.
        ini = valores.get("data inicio", [])
        fim = valores.get("data fim", [])
        if len(ini) < 2 or len(fim) < 2 or "data credito" not in valores:
            continue
        col["INICIO_FERIAS"] = ini[-1]
        col["FIM_FERIAS"] = fim[-1]
        col["CREDITO"] = valores["data credito"][0]

        if "dias" not in valores:
            continue
        col["DIAS"] = valores["dias"][-1]
        for chave, alvos in [("ABONO", ["qtd dias abono", "dias de abono"]),
                             ("DEC13", ["13o sal.", "13o sal", "13 sal."]),
                             ("SITUACAO", ["status"])]:
            for a in alvos:
                if a in valores:
                    col[chave] = valores[a][-1]
                    break
        return r, col
    return None


# ==========================================================================
# 2 e 3) Consulta Ficha (formato longo: uma linha por verba)
# ==========================================================================
def _ler_ficha_bruta(conteudo: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws = wb[wb.sheetnames[0]]
    cabecalho = [_sem_acento(_texto(c.value)) for c in ws[1]]

    def indice(*nomes):
        for n in nomes:
            if n in cabecalho:
                return cabecalho.index(n)
        return None

    i_mat, i_verba, i_valor = indice("matricula"), indice("verba"), indice("valor")
    if i_mat is None or i_verba is None or i_valor is None:
        raise FormatoInvalido(
            "Este arquivo não parece ser uma Consulta Ficha do sistema — esperava as "
            "colunas 'Matricula', 'Verba' e 'Valor' na primeira linha."
        )
    i_mes = indice("mes e ano")
    i_proc = indice("processo")
    i_nome = indice("nome funcionario")
    i_qtd = indice("quantidade")
    i_sal = indice("valor do salario")

    registros = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        mat = _matricula(r[i_mat]) if i_mat < len(r) else None
        if mat is None:
            continue
        registros.append({
            "MAT": mat,
            "MES": _texto(r[i_mes]) if i_mes is not None and i_mes < len(r) else "",
            "PROCESSO": _texto(r[i_proc]) if i_proc is not None and i_proc < len(r) else "",
            "NOME": _texto(r[i_nome]) if i_nome is not None and i_nome < len(r) else "",
            "VERBA": _matricula(r[i_verba]),
            "QTD": _num(r[i_qtd]) if i_qtd is not None and i_qtd < len(r) else 0.0,
            "VALOR": _num(r[i_valor]),
            "SALARIO_REF": _num(r[i_sal]) if i_sal is not None and i_sal < len(r) else 0.0,
        })
    if not registros:
        raise FormatoInvalido("A Consulta Ficha enviada está vazia.")
    return pd.DataFrame(registros)


def ler_salarios(conteudo: bytes) -> tuple:
    """Ficha do MÊS ANTERIOR: devolve (dict matricula -> salário 644, competência).

    Se o arquivo não tiver a verba 644 em nenhuma linha, avisa — quase sempre
    significa que foi exportado o mês errado ou sem essa verba."""
    df = _ler_ficha_bruta(conteudo)
    so644 = df[df["VERBA"] == VERBA_SALARIO]
    if so644.empty:
        raise FormatoInvalido(
            "Não achei a verba 644 (Salário Mês c/ Adic.) nesta ficha. É ela que "
            "define o salário usado no cálculo das férias — confira se exportou o "
            "mês anterior ao das férias e com essa verba incluída."
        )
    salarios = so644.groupby("MAT")["VALOR"].max().to_dict()
    competencias = sorted({m for m in so644["MES"] if m})
    return salarios, (competencias[0] if len(competencias) == 1 else ", ".join(competencias))


def ler_ficha_ferias(conteudo: bytes) -> tuple:
    """Ficha do MÊS DAS FÉRIAS: transforma o formato longo em uma linha por
    funcionário, com cada grupo de verbas somado na sua coluna.
    Devolve (dict matricula -> valores, competência)."""
    df = _ler_ficha_bruta(conteudo)
    ferias = df[df["PROCESSO"].str.lower().str.startswith("f")] if "PROCESSO" in df else df
    if ferias.empty:
        ferias = df

    dados = {}
    for mat, grupo in ferias.groupby("MAT"):
        por_verba = grupo.groupby("VERBA")["VALOR"].sum().to_dict()
        qtd_por_verba = grupo.groupby("VERBA")["QTD"].sum().to_dict()
        linha = {"NOME_FICHA": grupo["NOME"].iloc[0],
                 "SALARIO_REF": grupo["SALARIO_REF"].max()}

        for campo, codigos in VERBAS.items():
            # a primeira verba da tupla é a "(REC)" — o recibo fechado, que é o
            # que interessa para conferir. As demais ficam guardadas à parte.
            linha[campo] = float(por_verba.get(codigos[0], 0.0))
            linha[campo + "_total"] = float(sum(por_verba.get(c, 0.0) for c in codigos))

        # quando o recibo (REC) não veio, cai para a soma das competências
        for campo in VERBAS:
            if linha[campo] == 0.0 and linha[campo + "_total"] != 0.0:
                linha[campo] = linha[campo + "_total"]

        linha["inss_faixas"] = float(sum(por_verba.get(c, 0.0) for c in VERBAS_INSS_FAIXAS))
        linha["qtd_dias_ferias"] = float(qtd_por_verba.get(147, 0.0) or qtd_por_verba.get(148, 0.0))
        linha["salario_644"] = float(por_verba.get(VERBA_SALARIO, 0.0))
        dados[mat] = linha

    competencias = sorted({m for m in ferias["MES"] if m})
    return dados, (competencias[0] if len(competencias) == 1 else ", ".join(competencias))


def competencia_anterior(mes_ano: str) -> str:
    """'09/2026' -> '08/2026'. Usado para conferir se a ficha de salário enviada
    é mesmo a do mês anterior ao das férias."""
    try:
        m, a = mes_ano.split("/")
        m, a = int(m), int(a)
    except Exception:
        return ""
    return f"{12:02d}/{a-1}" if m == 1 else f"{m-1:02d}/{a}"
